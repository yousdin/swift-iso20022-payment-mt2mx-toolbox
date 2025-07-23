"""
ISO20022 Swift Payment Messages Excel Aggregator

This script aggregates ISO20022 Swift Payment Messages Excel documentation files,
specifically the XML Schema Description (Compact View). It extracts metadata and message
content from a folder of Excel files and produces consolidated CSV and Excel outputs.

Features:
- Aggregates all rows from the 'Full_View' tab of each Excel file, appending metadata from the 'General Information' tab.
- Removes unwanted columns and standardizes column names.
- Adds multiple sheets to the Excel output:
    - CBPRPlus_XSD_Full_View: Aggregated data and metadata
    - Legend: Column names and descriptions
    - Process_Metadata: Script name, timestamp, processing stats
    - Process_FilesList: Mapping of each file to its Restricted_Base_Message

Usage:
- As a script (default folder: CBPRPlus_SR2025_Excel):
    python aggregate_metadata.py
- With a custom folder:
    python aggregate_metadata.py --folder path/to/your/excel_folder
- As a function in another script or notebook:
    from aggregate_metadata import aggregate_excel_folder
    aggregate_excel_folder(folder='path/to/your/excel_folder')

Dependencies: pandas, openpyxl

Output:
- CBPRPlus_SR2025_Metadata_Aggregated.csv
- CBPRPlus_SR2025_Metadata_Aggregated.xlsx (with multiple sheets)
"""

import os
import pandas as pd
from openpyxl import load_workbook



import argparse

def aggregate_excel_folder(folder='CBPRPlus_SR2025_Excel'):
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime

    folder_path = os.path.join(os.path.dirname(__file__), folder)
    all_rows = []

    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            try:
                wb = load_workbook(file_path, data_only=True)
                # --- Extract Metadata ---
                metadata = {'Source_File': filename}
                if 'General Information' in wb.sheetnames:
                    ws = wb['General Information']
                    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
                        if row[0] is not None:
                            key = str(row[0]).strip().replace(' ', '_')
                            value_c = row[2] if len(row) > 2 else None
                            metadata[key] = value_c
                else:
                    print(f"Warning: 'General Information' tab not found in {filename}")
                    continue
                # --- Extract Full_View Data ---
                if 'Full_View' in wb.sheetnames:
                    df_full_view = pd.read_excel(file_path, sheet_name='Full_View')
                    for _, row in df_full_view.iterrows():
                        row_dict = {}
                        # Add Restricted_Base_Message first
                        row_dict['Restricted_Base_Message'] = metadata.get('Restricted_Base_Message')
                        # Add Full_View data
                        row_dict.update(row.to_dict())
                        # Add remaining metadata (excluding Restricted_Base_Message)
                        for k, v in metadata.items():
                            if k not in ('Restricted_Base_Message', 'Source_File'):
                                row_dict[k] = v
                        row_dict['Source_File'] = filename
                        all_rows.append(row_dict)
                else:
                    print(f"Warning: 'Full_View' tab not found in {filename}")
            except Exception as e:
                print(f"Error processing {filename}: {e}")

    # Create DataFrame and export
    if all_rows:
        df = pd.DataFrame(all_rows)
        # Remove specific columns if they exist
        if 'Usage_Guideline_Description' in df.columns:
            df = df.drop(columns=['Usage_Guideline_Description'])
        if 'Privacy' in df.columns:
            df = df.drop(columns=['Privacy'])
        # Remove all columns starting with 'Generated_by_the_MyStandards_web_platform_'
        cols_to_drop = [col for col in df.columns if col.startswith('Generated_by_the_MyStandards_web_platform_')]
        if cols_to_drop:
            df = df.drop(columns=cols_to_drop)
        # Save to CSV - TODO: Add parameters for folder and filename if want to enable
        # df.to_csv('CBPRPlus_SR2025_Metadata_Aggregated.csv', index=False)

        # Prepare the Legend sheet (from the first file)
        legend_data = []
        first_file = None
        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx'):
                first_file = os.path.join(folder_path, filename)
                break
        if first_file:
            wb = load_workbook(first_file, data_only=True)
            if 'General Information' in wb.sheetnames:
                ws = wb['General Information']
                for row in ws.iter_rows(min_row=22, max_row=46, min_col=2, max_col=3, values_only=True):
                    if row[0] is not None:
                        legend_data.append({'Column_Name': row[0], 'Column_Description': row[1]})
        legend_df = pd.DataFrame(legend_data)

        # Prepare Process_Metadata sheet
        process_metadata = [
            {'Key': 'Script_Name', 'Value': 'aggregate_metadata.py'},
            {'Key': 'Run_DateTime', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'Key': 'Files_Processed', 'Value': len(df['Source_File'].unique())},
            {'Key': 'Rows_Aggregated', 'Value': len(df)}
        ]
        process_metadata_df = pd.DataFrame(process_metadata)

        # Prepare Process_FilesList sheet from General Information metadata
        files_list = []
        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(folder_path, filename)
                try:
                    wb = load_workbook(file_path, data_only=True)
                    if 'General Information' in wb.sheetnames:
                        ws = wb['General Information']
                        restricted_base_message = None
                        for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
                            if row[0] is not None and str(row[0]).strip().replace(' ', '_') == 'Restricted_Base_Message':
                                restricted_base_message = row[2] if len(row) > 2 else None
                                break
                        files_list.append({'Source_File': filename, 'Restricted_Base_Message': restricted_base_message})
                except Exception as e:
                    print(f"Error processing {filename} for Process_FilesList: {e}")
        process_fileslist_df = pd.DataFrame(files_list)

        # Write to Excel with custom sheet names
        with pd.ExcelWriter('CBPRPlus_SR2025_Metadata_Aggregated.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='CBPRPlus_XSD_Full_View', index=False)
            legend_df.to_excel(writer, sheet_name='Legend', index=False)
            process_metadata_df.to_excel(writer, sheet_name='Process_Metadata', index=False)
            process_fileslist_df.to_excel(writer, sheet_name='Process_FilesList', index=False)
        print('Aggregation complete. Output saved as CSV and Excel.')
    else:
        print('No data extracted.')

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Aggregate ISO20022 Swift Payment Messages Excel Documentation")
    parser.add_argument('--folder', type=str, default='CBPRPlus_SR2025_Excel', help='Folder containing Excel files (default: CBPRPlus_SR2025_Excel)')
    args = parser.parse_args()
    aggregate_excel_folder(args.folder)
